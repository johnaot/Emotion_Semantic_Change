"""
Useful functions
"""
import numpy as np
import gensim
import pandas as pd
from openpyxl import Workbook, load_workbook

import string
import pickle, os
from collections import defaultdict

from constants import *

def similarity(a, b):
  return np.dot(a,b)/(np.linalg.norm(a)*np.linalg.norm(b))

def distance(a, b):
  return 1 - similarity(a,b)


#################### Read data form french list 1998 ##################
'''
get french words from appendix
'''
def read_table_1998(p):
  data = defaultdict(int)
  words, ratings = [], []
  with open(p + '/french_edit', 'r') as f:
    lines = f.readlines()
    words = [line.strip() for line in lines if line != '\n']
  with open(p + '/ratings', 'r') as f:
    content = f.read()
    content = content.replace('\n', ' ')[0:-1]
    ratings = content.split(' ')
  for i, w in enumerate(words):
    data[w.lower()] = float(ratings[i])
  return data

def read_table_1998_valence(p):
  data = defaultdict(int)
  words, valence = [], []
  with open(p + '/french_edit', 'r') as f:
    lines = f.readlines()
    words = [line.strip() for line in lines if line != '\n']
  with open(p + '/valence', 'r') as f:
    lines = f.readlines()
    valence = [v.strip() for v in lines]
  for i, w in enumerate(words):
    data[w.lower()] = float(valence[i])
  return data

'''
get english equivalents
'''
def read_english_1998(p):
  data = defaultdict(int)
  words, eng_words = [], []
  with open(p + '/french_edit', 'r') as f:
    lines = f.readlines()
    words = [line.strip() for line in lines if line != '\n']
  with open(p + '/english', 'r') as f:
    lines = f.readlines()
    eng_words = [line.strip() for line in lines if line != '\n']
  for i, w in enumerate(words):
    data[w.lower()] = eng_words[i].lower()
  return data

#################### Read data form 1987 list ####################
'''
get emotion words from Table 1
'''
def read_table_1987(p):
  data = defaultdict(int)
  with open(p, 'r') as f:
    lines = f.readlines()
    for line in lines:
      line = line.split(' ')
      if len(line) == 1:
        continue
      data[line[0]] = float(line[1])
  return data

def read_table2_1987(p):
  data = defaultdict(int)
  with open(p, 'r') as f:
    lines = f.readlines()
    i = 0
    words = []
    while i < len(lines):
      line = lines[i].strip()
      if line == 'Evaluation':
        for w in words:
          i += 1
          line = lines[i].strip()
          data[w] = float(line)
      elif line == 'Intensity':
        for w in words:
          i += 1
        words = []
      else:
        words.append(line)
      i += 1
  return data

#################### Read data form 1980 norms ####################
'''
get words from 1980 norms
'''
def read_norms_1980(p):
  data = {}
  is_name = lambda w: w.replace('-', '').isalpha()
  with open(p, 'r') as file:
    lines = file.readlines()
    current_domain = ''
    for i, line in enumerate(lines):
      items = line.strip().split(' ')
      if i % 4 == 0:
        current_domain = line.strip()
      elif i % 4 == 2:
        data[current_domain] = {}
        items2 = []
        for item in items:
          if is_name(item) and len(items2) > 0 and is_name(items2[-1]):
            items2[-1] = items2[-1] + '-' + item
          else:
            items2.append(item)
        for i in range(0,len(items2),3): 
          data[current_domain][items2[i].lower()] = float(items2[i+1]) 
  return data

#################### Read data form Leuven ####################
'''
get words from Leuven
'''
def read_leuven(p):
  sheetnames = ['birds', 'fish', 'insects', 'mammals', 'reptiles',]
  data = defaultdict(dict)
  for sheetname in sheetnames:
    df = pd.read_excel(p, sheet_name=sheetname)
    words = list(df['Unnamed: 1'])
    values = list(df['mean'])
    for i, w in enumerate(words):
      data[sheetname][w] = values[i]
  return data

#################### Read data form Hamilton ####################
def get_hamilton_data(time_start, time_end, time_delta):
  time_range = list(range(time_start, time_end, time_delta))

  npy_path = os.path.join(npy_path_eng, "%d-w.npy")
  vocab_path = os.path.join(vocab_path_eng, "%d-vocab.pkl")
  pos_path = os.path.join(pos_path_eng, "%d-pos.pkl")

  data = defaultdict(list)
  pos_data = defaultdict(list)
  for t in time_range:
    vocab = pickle.load(open(vocab_path % t, 'rb'), encoding='latin1')
    pos = pickle.load(open(pos_path % t, 'rb'), encoding='latin1')
    vectors = np.load(npy_path % t)

    for i, w in enumerate(vocab):
      data[w].append(vectors[i])
      if w in pos:
        if pos[w] == 'NOUN':
          pos_data[w].append(NOUN)
        elif pos[w] == 'ADJ':
          pos_data[w].append(ADJ)
        elif pos[w] == 'VERB':
          pos_data[w].append(VERB)
        else:
          pos_data[w].append(pos_placeholder)
      else:
        pos_data[w].append(pos_placeholder)

  return(data, pos_data)

def get_hamilton_data_french(time_start, time_end, time_delta):
  time_range = list(range(time_start, time_end, time_delta))

  npy_path = os.path.join(npy_path_fra, "%d-w.npy")
  vocab_path = os.path.join(vocab_path_fra, "%d-vocab.pkl")
  pos_path = os.path.join(pos_path_fra, "%d-pos.pkl")

  data = defaultdict(list)
  pos_data = defaultdict(list)
  freq_data = pickle.load(open(freq_path_fra, 'rb'), encoding='latin1')
  for t in time_range:
    vocab = pickle.load(open(vocab_path % t, 'rb'), encoding='latin1')
    pos = pickle.load(open(pos_path % t, 'rb'), encoding='latin1')
    vectors = np.load(npy_path % t)

    for i, w in enumerate(vocab):
      data[w].append(vectors[i])
      if w in pos:
        if pos[w] == 'NOUN':
          pos_data[w].append(NOUN)
        elif pos[w] == 'ADJ':
          pos_data[w].append(ADJ)
        elif pos[w] == 'VERB':
          pos_data[w].append(VERB)
        else:
          pos_data[w].append(pos_placeholder)
      else:
        pos_data[w].append(pos_placeholder)

  return(data, pos_data, freq_data)
